"""모든 리스트 엑셀 다운로드 e2e."""
import uuid as _uuid
from io import BytesIO

from openpyxl import load_workbook

from app.admin.export import XLSX_MEDIA
from tests.factories import create_plan, create_service, create_subscription, create_user
from tests.helpers import admin_login


def _wb(resp):
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MEDIA
    assert "attachment" in resp.headers["content-disposition"]
    return load_workbook(BytesIO(resp.content)).active


async def test_services_export(client, db, redis_client, cipher):
    await create_service(db, cipher, name="엑셀서비스A")
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    ws = _wb(await client.get("/admin/services/export.xlsx"))
    assert [c.value for c in ws[1]] == ["서비스명", "담당자 이메일", "허용 IP", "상태"]
    names = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert "엑셀서비스A" in names


async def test_users_export(client, db, redis_client, cipher):
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    ws = _wb(await client.get("/admin/users/export.xlsx"))
    assert [c.value for c in ws[1]] == ["이메일", "역할", "주 서비스", "상태"]
    assert any(admin.email == row[0].value for row in ws.iter_rows(min_row=2))


async def test_plans_export(client, db, redis_client, cipher):
    svc, _, _ = await create_service(db, cipher, name="요금제서비스")
    await create_plan(db, svc, name="베이직요금")
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    ws = _wb(await client.get("/admin/plans/export.xlsx"))
    assert [c.value for c in ws[1]] == ["서비스", "요금제", "결제주기", "정가",
                                        "첫 결제", "정기 결제", "상태"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(r[1] == "베이직요금" for r in rows)


async def test_subscriptions_export(client, db, redis_client, cipher):
    svc, _, _ = await create_service(db, cipher, name="구독서비스")
    plan = await create_plan(db, svc)
    await create_subscription(db, cipher, svc, plan, external_user_id="exp-user@e.com")
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    ws = _wb(await client.get("/admin/subscriptions/export.xlsx"))
    assert [c.value for c in ws[1]] == ["서비스", "사용자", "요금제", "상태",
                                        "만료일", "다음 결제"]
    assert any(r[1] == "exp-user@e.com" for r in ws.iter_rows(min_row=2, values_only=True))


async def test_payments_export_scoped_to_manager(client, db, redis_client, cipher):
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc_a, _, _ = await create_service(db, cipher, name="결제A")
    svc_b, _, _ = await create_service(db, cipher, name="결제B")
    for svc, oid in [(svc_a, "exp-a"), (svc_b, "exp-b")]:
        db.add(Payment(subscription_id=None, service_id=svc.id, external_user_id="u@e.com",
                       order_id=oid, amount=1000, payment_type=PaymentType.ONE_OFF,
                       kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                       idempotency_key=oid, requested_at=utcnow(), approved_at=utcnow()))
    await db.commit()
    mgr, pw = await create_user(db, role="SERVICE_MANAGER", service_id=svc_a.id)
    await admin_login(client, mgr.email, pw)
    ws = _wb(await client.get("/admin/payments/export.xlsx"))
    assert [c.value for c in ws[1]] == ["주문번호", "서비스", "종류", "사용자", "유형",
                                        "금액", "상태", "실패코드", "요청시각"]
    orders = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)]
    assert "exp-a" in orders and "exp-b" not in orders   # 매니저 스코프 격리


async def test_settlement_export_manager_other_service_404(client, db, redis_client, cipher):
    svc_a, _, _ = await create_service(db, cipher, name="정산매니저A")
    svc_b, _, _ = await create_service(db, cipher, name="정산매니저B")
    mgr, pw = await create_user(db, role="SERVICE_MANAGER", service_id=svc_a.id)
    await admin_login(client, mgr.email, pw)
    resp = await client.get(f"/admin/settlement/export.xlsx?service_id={svc_b.id}")
    assert resp.status_code == 404


async def test_settlement_export_all_mode(client, db, redis_client, cipher):
    from datetime import datetime, timezone
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc, _, _ = await create_service(db, cipher, name="정산서비스")
    when = datetime(2026, 6, 3, tzinfo=timezone.utc)
    db.add(Payment(subscription_id=None, service_id=svc.id, external_user_id="u@e.com",
                   order_id="set-oo", amount=5000, payment_type=PaymentType.ONE_OFF,
                   kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                   idempotency_key="set-oo", requested_at=when, approved_at=when))
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    ws = _wb(await client.get(
        "/admin/settlement/export.xlsx?from=2026-06-01&to=2026-06-30"))
    assert [c.value for c in ws[1]] == ["서비스", "건수", "구독매출", "일반매출",
                                        "총매출", "환불", "순매출"]
    assert any(r[0] == "정산서비스" for r in ws.iter_rows(min_row=2, values_only=True))


async def test_settlement_export_service_mode(client, db, redis_client, cipher):
    from datetime import datetime, timezone
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc, _, _ = await create_service(db, cipher, name="정산상세")
    when = datetime(2026, 6, 3, tzinfo=timezone.utc)
    db.add(Payment(subscription_id=None, service_id=svc.id, external_user_id="su@e.com",
                   order_id="set-detail", amount=7000, payment_type=PaymentType.ONE_OFF,
                   kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                   idempotency_key="set-detail", requested_at=when, approved_at=when))
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    ws = _wb(await client.get(
        f"/admin/settlement/export.xlsx?from=2026-06-01&to=2026-06-30&service_id={svc.id}"))
    assert [c.value for c in ws[1]] == ["승인시각", "사용자", "주문번호", "유형",
                                        "종류", "상태", "총매출", "환불", "순매출"]
    assert any(r[2] == "set-detail" for r in ws.iter_rows(min_row=2, values_only=True))


async def test_service_detail_exports(client, db, redis_client, cipher):
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc, _, _ = await create_service(db, cipher, name="상세서비스")
    plan = await create_plan(db, svc, name="상세요금")
    await create_subscription(db, cipher, svc, plan, external_user_id="d-sub@e.com")
    db.add(Payment(subscription_id=None, service_id=svc.id, external_user_id="d-oo@e.com",
                   order_id="d-oo-1", amount=3000, payment_type=PaymentType.ONE_OFF,
                   kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                   idempotency_key="d-oo-1", requested_at=utcnow(), approved_at=utcnow()))
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    subs = _wb(await client.get(f"/admin/services/{svc.id}/subs.xlsx"))
    assert [c.value for c in subs[1]] == ["사용자", "요금제", "상태", "만료일", "다음 결제"]
    assert any(r[0] == "d-sub@e.com" for r in subs.iter_rows(min_row=2, values_only=True))
    oneoff = _wb(await client.get(f"/admin/services/{svc.id}/oneoff.xlsx"))
    assert [c.value for c in oneoff[1]] == ["승인시각", "사용자", "주문번호", "금액",
                                            "환불", "수수료", "상태"]
    assert any(r[2] == "d-oo-1" for r in oneoff.iter_rows(min_row=2, values_only=True))
    plans = _wb(await client.get(f"/admin/services/{svc.id}/plans.xlsx"))
    assert [c.value for c in plans[1]] == ["요금제", "결제주기", "정가", "첫 결제",
                                           "정기 결제", "상태"]
    assert any(r[0] == "상세요금" for r in plans.iter_rows(min_row=2, values_only=True))


async def test_service_detail_export_404_for_unknown_service(client, db, redis_client, cipher):
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    missing = _uuid.uuid4()
    for suffix in ("subs", "oneoff", "plans"):
        resp = await client.get(f"/admin/services/{missing}/{suffix}.xlsx")
        assert resp.status_code == 404


async def test_payments_export_reflects_status_filter(client, db, redis_client, cipher):
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc, _, _ = await create_service(db, cipher, name="필터서비스")
    for oid, st in [("flt-done", PaymentStatus.DONE), ("flt-fail", PaymentStatus.FAILED)]:
        db.add(Payment(subscription_id=None, service_id=svc.id, external_user_id="u@e.com",
                       order_id=oid, amount=1000, payment_type=PaymentType.ONE_OFF,
                       kind=PaymentKind.ONE_OFF, status=st,
                       idempotency_key=oid, requested_at=utcnow(),
                       approved_at=(utcnow() if st == PaymentStatus.DONE else None)))
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    ws = _wb(await client.get("/admin/payments/export.xlsx?status=FAILED"))
    orders = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)]
    assert orders == ["flt-fail"]      # status 필터로 1건만


async def test_list_pages_show_export_buttons(client, db, redis_client, cipher):
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    checks = {
        "/admin/services": "/admin/services/export.xlsx",
        "/admin/plans": "/admin/plans/export.xlsx",
        "/admin/subscriptions": "/admin/subscriptions/export.xlsx",
        "/admin/payments": "/admin/payments/export.xlsx",
        "/admin/users": "/admin/users/export.xlsx",
    }
    for page, link in checks.items():
        html = (await client.get(page)).text
        assert link in html, f"{page}에 export 버튼 없음"
