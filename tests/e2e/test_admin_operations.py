import json

from sqlalchemy import select

from app.models import AuditLog, PasswordSetupToken
from tests.factories import create_plan, create_service, create_subscription, create_user
from tests.helpers import admin_login, api_request, get_csrf


async def test_saved_query_emits_hx_trigger(client, db, redis_client, cipher):
    """?saved= 가 있는 어드민 페이지는 HX-Trigger:{"showSaved":..} 헤더를 보낸다 (요청).

    htmx로 쓰는 액션(스왑)에서도 완료 모달이 뜨도록 하는 경로 — 헤더로 검증.
    일반 전체 페이지 로드에서는 body[data-saved]가 모달을 띄운다.
    """
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    resp = await client.get("/admin/settings?saved=저장되었습니다")
    assert resp.status_code == 200
    trigger = resp.headers.get("HX-Trigger")
    assert trigger and json.loads(trigger).get("showSaved") == "저장되었습니다"
    # body[data-saved]도 함께 세팅되어 비-htmx 로드에서 모달이 뜬다
    assert 'data-saved="저장되었습니다"' in resp.text


async def test_settings_page_and_forms(client, db, redis_client, cipher):
    """전체설정 화면 렌더 + 재시도 폼 저장 + 킬스위치 ON 검증 (요청 013).

    1. SYSTEM_ADMIN으로 로그인 후 GET /admin/settings → 200.
    2. 재시도 설정 저장 → 303 redirect with ?saved= + gs.retry_limit 반영 확인.
    3. 킬스위치 ON (비번 필요) → 303 redirect with ?saved= + gs.server_disabled=True 확인.

    성공 redirect 검증: follow_redirects=False 기준, Location 헤더에 "saved" 포함 여부.
    """
    from app.services import app_settings

    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    session_id = await admin_login(client, admin.email, pw)

    # 설정 페이지 렌더
    r = await client.get("/admin/settings")
    assert r.status_code == 200, f"GET /admin/settings: {r.status_code}"

    csrf = await get_csrf(redis_client, session_id)

    # 재시도 설정 저장: 303 redirect + Location에 ?saved= 포함 확인(완료 모달 트리거)
    r = await client.post(
        "/admin/settings/retry",
        data={
            "csrf_token": csrf,
            "retry_limit": "5",
            "retry_interval_hours": "8",
            "suspended_grace_days": "20",
        },
        follow_redirects=False,
    )
    assert r.status_code == 303, f"POST retry: expected 303, got {r.status_code}"
    assert "saved" in r.headers.get("location", ""), (
        f"POST retry: Location should contain 'saved', got {r.headers.get('location')}"
    )
    db.expire_all()
    gs = await app_settings.get_global_settings(db)
    assert gs.retry_limit == 5, f"retry_limit expected 5, got {gs.retry_limit}"

    # 킬스위치 ON — 본인 비밀번호 필요; 303 redirect + Location에 saved 포함 확인
    r = await client.post(
        "/admin/settings/server-toggle",
        data={
            "csrf_token": csrf,
            "disabled": "on",
            "reason": "점검",
            "password": pw,
        },
        follow_redirects=False,
    )
    assert r.status_code == 303, f"POST server-toggle: expected 303, got {r.status_code}"
    assert "saved" in r.headers.get("location", ""), (
        f"POST server-toggle: Location should contain 'saved', got {r.headers.get('location')}"
    )
    db.expire_all()
    gs2 = await app_settings.get_global_settings(db)
    assert gs2.server_disabled is True, "server_disabled should be True after toggle ON"


async def test_settings_security_policy_form(client, db, redis_client, cipher):
    """전체설정 '보안/결제 정책' 폼 저장 → 303 ?saved= + GlobalSettings 반영(런타임 적용)."""
    from app.services import app_settings

    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    session_id = await admin_login(client, admin.email, pw)
    # 폼 필드가 화면에 노출되는지
    page = (await client.get("/admin/settings")).text
    assert "보안/결제 정책" in page and 'name="one_off_max_amount"' in page
    csrf = await get_csrf(redis_client, session_id)
    r = await client.post(
        "/admin/settings/security-policy",
        data={
            "csrf_token": csrf,
            "max_failed_logins": "3",
            "account_lock_minutes": "10",
            "one_off_max_amount": "50000",
        },
        follow_redirects=False,
    )
    assert r.status_code == 303
    assert "saved" in r.headers.get("location", "")
    db.expire_all()
    gs = await app_settings.get_global_settings(db)
    assert gs.max_failed_logins == 3
    assert gs.account_lock_minutes == 10
    assert gs.one_off_max_amount == 50000


async def test_settings_admin_ips_form(client, db, redis_client, cipher):
    """admin-ips 폼 저장 e2e: 루프백(127.0.0.1)은 입력해도 저장에서 자동 제외되고,
    나머지 IP만 정상 저장된다.

    127.0.0.1은 같은 서버라 항상 허용이므로 목록에 보관하지 않는다. 테스트 클라이언트
    IP(127.0.0.1)는 루프백이라 스스로 잠금에 빠지지 않으므로, lockout 거부는
    별도 비루프백 테스트(test_admin_ips_lockout_blocks_self_exclusion)에서 검증한다.
    폼 필드: admin_allowed_ips (줄바꿈 구분), POST /admin/settings/admin-ips.
    """
    from app.services import app_settings

    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    session_id = await admin_login(client, admin.email, pw)
    csrf = await get_csrf(redis_client, session_id)

    # 127.0.0.1 + 10.0.0.2 저장 → 성공. 루프백은 제외되고 10.0.0.2만 저장된다.
    r = await client.post(
        "/admin/settings/admin-ips",
        data={"csrf_token": csrf, "admin_allowed_ips": "127.0.0.1\n10.0.0.2"},
        follow_redirects=False,
    )
    assert r.status_code == 303, (
        f"POST admin-ips(허용): expected 303, got {r.status_code}"
    )
    assert "saved" in r.headers.get("location", ""), (
        f"POST admin-ips(허용): Location에 'saved' 없음: {r.headers.get('location')}"
    )
    # DB 반영 확인 — 루프백은 빠지고 10.0.0.2만 저장
    db.expire_all()
    gs = await app_settings.get_global_settings(db)
    assert "127.0.0.1" not in gs.admin_allowed_ips, (
        f"루프백은 저장되지 않아야 함: {gs.admin_allowed_ips}"
    )
    assert gs.admin_allowed_ips == ["10.0.0.2"], gs.admin_allowed_ips


async def test_admin_ips_lockout_blocks_self_exclusion(app, db, redis_client, cipher):
    """비루프백 어드민이 자기 IP를 뺀 목록을 저장하려 하면 lockout 거부(?error=)."""
    from app.services import app_settings
    from tests.helpers import client_from_ip

    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    # 현재 비루프백 IP를 허용해 어드민 접근이 가능하게 둔다
    gs = await app_settings.get_global_settings(db)
    gs.admin_allowed_ips = ["203.0.113.50"]
    await db.commit()

    async with client_from_ip(app, "203.0.113.50") as ext:
        session_id = await admin_login(ext, admin.email, pw)
        csrf = await get_csrf(redis_client, session_id)
        # 자기 IP(203.0.113.50)를 뺀 목록 저장 시도 → lockout 거부
        r = await ext.post(
            "/admin/settings/admin-ips",
            data={"csrf_token": csrf, "admin_allowed_ips": "10.0.0.99"},
            follow_redirects=False,
        )
    assert r.status_code == 303
    location = r.headers.get("location", "")
    assert "error" in location and "saved" not in location, location


async def test_settings_page_forbidden_for_manager(client, db, redis_client, cipher):
    """SERVICE_MANAGER는 /admin/settings에 접근 불가(require_admin=SYSTEM_ADMIN 전용)."""
    svc, _, _ = await create_service(db, cipher)
    manager, pw = await create_user(db, role="SERVICE_MANAGER", service_id=svc.id)
    await admin_login(client, manager.email, pw)
    r = await client.get("/admin/settings")
    assert r.status_code == 403, f"Expected 403 for SERVICE_MANAGER, got {r.status_code}"


async def test_admin_ip_restriction(client, app, db, redis_client, cipher):
    """어드민 IP 제한: 허용목록에 없는 비루프백 IP는 403, 루프백(127.0.0.1)은 항상 허용.

    127.0.0.1/::1(같은 서버)은 화이트리스트와 무관하게 통과하므로, 제한 차단은
    비루프백 소스 IP(client_from_ip)로 검증한다.
    """
    from app.services import app_settings
    from tests.helpers import client_from_ip
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")

    # 허용 목록을 비루프백 IP 1개로 제한
    gs = await app_settings.get_global_settings(db)
    gs.admin_allowed_ips = ["10.0.0.1"]
    await db.commit()

    # ① 목록에 없는 비루프백 IP(203.0.113.9) → 403
    async with client_from_ip(app, "203.0.113.9") as ext:
        await admin_login(ext, admin.email, pw)   # 로그인은 IP 제한 없음
        r = await ext.get("/admin/services")
        assert r.status_code == 403

    # ② 루프백(127.0.0.1)은 제한 목록이 있어도 항상 허용 → 200
    await admin_login(client, admin.email, pw)
    r2 = await client.get("/admin/services")
    assert r2.status_code == 200


async def test_manager_sees_only_own_subscriptions(client, db, redis_client, cipher):
    svc_a, _, _ = await create_service(db, cipher, name="sub-svc-a")
    svc_b, _, _ = await create_service(db, cipher, name="sub-svc-b")
    plan_a = await create_plan(db, svc_a)
    plan_b = await create_plan(db, svc_b)
    await create_subscription(db, cipher, svc_a, plan_a, external_user_id="user-of-a")
    await create_subscription(db, cipher, svc_b, plan_b, external_user_id="user-of-b")
    user, pw = await create_user(db, role="SERVICE_MANAGER", service_id=svc_a.id)
    await admin_login(client, user.email, pw)

    resp = await client.get("/admin/subscriptions")
    assert "user-of-a" in resp.text
    assert "user-of-b" not in resp.text


async def test_admin_sees_all_subscriptions(client, db, redis_client, cipher):
    svc, _, _ = await create_service(db, cipher)
    plan = await create_plan(db, svc)
    await create_subscription(db, cipher, svc, plan, external_user_id="user-all")
    user, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, user.email, pw)
    resp = await client.get("/admin/subscriptions")
    assert "user-all" in resp.text


async def test_manager_cannot_open_other_service_subscription_detail(
        client, db, redis_client, cipher):
    svc_a, _, _ = await create_service(db, cipher, name="det-svc-a")
    svc_b, _, _ = await create_service(db, cipher, name="det-svc-b")
    plan_b = await create_plan(db, svc_b)
    sub_b = await create_subscription(db, cipher, svc_b, plan_b)
    user, pw = await create_user(db, role="SERVICE_MANAGER", service_id=svc_a.id)
    await admin_login(client, user.email, pw)
    resp = await client.get(f"/admin/subscriptions/{sub_b.id}")
    assert resp.status_code == 404


async def test_subscription_detail_shows_trial_in_progress(client, db, redis_client, cipher):
    """체험 중(TRIAL) 구독 상세에 체험 정보(기간·진행 상태)가 노출된다."""
    svc, _, _ = await create_service(db, cipher)
    plan = await create_plan(db, svc, trial_enabled=True, trial_days=14)
    sub = await create_subscription(db, cipher, svc, plan, external_user_id="u-trial",
                                    status="TRIAL")
    user, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, user.email, pw)
    html = (await client.get(f"/admin/subscriptions/{sub.id}")).text
    assert "체험" in html
    assert "체험 중" in html
    assert "14일" in html


async def test_subscription_detail_shows_no_trial_for_plain_plan(client, db, redis_client, cipher):
    """체험 미제공 요금제 구독 상세에는 '체험 미제공'으로 표기된다."""
    svc, _, _ = await create_service(db, cipher)
    plan = await create_plan(db, svc)  # trial_enabled=False
    sub = await create_subscription(db, cipher, svc, plan, external_user_id="u-notrial")
    user, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, user.email, pw)
    html = (await client.get(f"/admin/subscriptions/{sub.id}")).text
    assert "체험 미제공" in html


async def test_subscription_detail_shows_trial_used_after_conversion(
        client, db, redis_client, cipher):
    """체험 후 전환된(ACTIVE) 구독도 가입 시 체험 사용 이력(가입 감사로그 기반)을 노출한다."""
    from app.services.audit import record_audit
    svc, _, _ = await create_service(db, cipher)
    plan = await create_plan(db, svc, trial_enabled=True, trial_days=7)
    sub = await create_subscription(db, cipher, svc, plan, external_user_id="u-converted",
                                    status="ACTIVE")
    await record_audit(db, actor_type="SERVICE", actor_service_id=svc.id,
                       action="subscription.create", target_type="subscription",
                       target_id=str(sub.id), detail={"trial": True})
    await db.commit()
    user, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, user.email, pw)
    html = (await client.get(f"/admin/subscriptions/{sub.id}")).text
    assert "체험 사용함" in html


async def test_force_cancel_subscription(client, db, redis_client, cipher):
    svc, _, _ = await create_service(db, cipher)
    plan = await create_plan(db, svc)
    sub = await create_subscription(db, cipher, svc, plan, external_user_id="u-force")
    user, pw = await create_user(db, role="SYSTEM_ADMIN")
    session_id = await admin_login(client, user.email, pw)
    csrf = await get_csrf(redis_client, session_id)

    resp = await client.post(f"/admin/subscriptions/{sub.id}/force-cancel",
                             data={"csrf_token": csrf})
    assert resp.status_code == 303
    await db.refresh(sub)
    assert sub.status == "CANCELED"
    log = await db.scalar(select(AuditLog).where(
        AuditLog.action == "subscription.force_cancel"))
    assert log is not None
    assert log.actor_user_id == user.id


async def test_payments_page_scoped(client, db, redis_client, cipher, fake_toss):
    svc, _, _ = await create_service(db, cipher)
    plan = await create_plan(db, svc)
    sub = await create_subscription(db, cipher, svc, plan, external_user_id="u-paylist")
    from app.models import Payment
    from app.core.clock import utcnow
    db.add(Payment(subscription_id=sub.id, order_id="adm-pay-1", amount=9900,
                   payment_type="RENEWAL", status="DONE", idempotency_key="ik",
                   requested_at=utcnow(),
                   service_id=sub.service_id, external_user_id=sub.external_user_id))
    await db.commit()
    user, pw = await create_user(db, role="SERVICE_MANAGER", service_id=svc.id)
    await admin_login(client, user.email, pw)
    resp = await client.get("/admin/payments")
    assert resp.status_code == 200
    assert "adm-pay-1" in resp.text


async def test_users_page_admin_only_and_reset_password(client, db, redis_client,
                                                        cipher, email_sender):
    svc, _, _ = await create_service(db, cipher)
    manager, _ = await create_user(db, role="SERVICE_MANAGER", service_id=svc.id)
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    session_id = await admin_login(client, admin.email, pw)
    csrf = await get_csrf(redis_client, session_id)

    page = await client.get("/admin/users")
    assert page.status_code == 200
    assert manager.email in page.text

    resp = await client.post(f"/admin/users/{manager.id}/reset-password",
                             data={"csrf_token": csrf})
    assert resp.status_code == 303
    token = await db.scalar(select(PasswordSetupToken).where(
        PasswordSetupToken.user_id == manager.id))
    assert token is not None
    assert any("비밀번호" in m["subject"] for m in email_sender.sent)


async def test_audit_page_lists_actions(client, db, redis_client, cipher):
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)  # auth.login 감사 로그 생성됨
    resp = await client.get("/admin/audit")
    assert resp.status_code == 200
    # 가독화: 액션 코드 대신 한글 라벨 + 행위자 이메일
    assert "로그인" in resp.text
    assert admin.email in resp.text


async def test_audit_page_forbidden_for_manager(client, db, redis_client, cipher):
    svc, _, _ = await create_service(db, cipher)
    user, pw = await create_user(db, role="SERVICE_MANAGER", service_id=svc.id)
    await admin_login(client, user.email, pw)
    resp = await client.get("/admin/audit")
    assert resp.status_code == 403


async def test_manager_cannot_force_cancel_other_service_subscription(
        client, db, redis_client, cipher):
    """매니저가 다른 서비스 구독을 강제취소 POST해도 404 (서비스 계층 격리)."""
    svc_a, _, _ = await create_service(db, cipher, name="fc-own")
    svc_b, _, _ = await create_service(db, cipher, name="fc-other")
    plan_b = await create_plan(db, svc_b)
    sub_b = await create_subscription(db, cipher, svc_b, plan_b, external_user_id="u-fcb")
    user, pw = await create_user(db, role="SERVICE_MANAGER", service_id=svc_a.id)
    session_id = await admin_login(client, user.email, pw)
    csrf = await get_csrf(redis_client, session_id)
    resp = await client.post(f"/admin/subscriptions/{sub_b.id}/force-cancel",
                             data={"csrf_token": csrf})
    assert resp.status_code == 404
    await db.refresh(sub_b)
    assert sub_b.status == "ACTIVE"


async def test_reset_password_destroys_target_user_sessions(client, db, redis_client, cipher):
    """관리자 비밀번호 재설정 발급 즉시 대상 사용자의 기존 세션 파기."""
    svc, _, _ = await create_service(db, cipher)
    manager, mpw = await create_user(db, role="SERVICE_MANAGER", service_id=svc.id)
    # 별도 클라이언트로 매니저 로그인
    from httpx import ASGITransport, AsyncClient
    async with AsyncClient(transport=ASGITransport(app=client._transport.app),
                           base_url="http://test") as mgr_client:
        await admin_login(mgr_client, manager.email, mpw)
        assert (await mgr_client.get("/admin")).status_code == 200

        admin, apw = await create_user(db, role="SYSTEM_ADMIN")
        session_id = await admin_login(client, admin.email, apw)
        csrf = await get_csrf(redis_client, session_id)
        resp = await client.post(f"/admin/users/{manager.id}/reset-password",
                                 data={"csrf_token": csrf})
        assert resp.status_code == 303
        # 매니저의 기존 세션은 무효화됨
        assert (await mgr_client.get("/admin")).status_code == 303


async def test_audit_resolves_target_and_detail(client, db, redis_client, cipher,
                                                email_sender):
    """감사로그가 대상 이름·상세를 사람이 읽을 수 있게 표시."""
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    sid = await admin_login(client, admin.email, pw)
    csrf = await get_csrf(redis_client, sid)
    mgr, _ = await create_user(db, role="SERVICE_MANAGER", service_id=None)
    # 서비스 등록 → service.register 감사(detail.name) 생성
    await client.post("/admin/services", data={
        "csrf_token": csrf, "name": "감사확인서비스",
        "manager_ids": [str(mgr.id)], "primary_user_id": str(mgr.id),
        "allowed_ips": "10.0.0.1"})
    resp = await client.get("/admin/audit")
    assert "서비스 등록" in resp.text           # 액션 한글
    assert "서비스 · 감사확인서비스" in resp.text  # 대상 타입+이름
    assert "이름 감사확인서비스" in resp.text     # detail 요약


async def test_audit_service_actor_shows_name_link(client, db, redis_client, cipher):
    """SERVICE 행위자 로그는 '외부 서비스 (서비스명)' + 상세 링크로 표시."""
    from app.services.audit import record_audit
    svc, _, _ = await create_service(db, cipher, name="행위자서비스")
    await record_audit(db, actor_type="SERVICE", actor_service_id=svc.id,
                       action="subscription.create", target_type="subscription",
                       target_id="00000000-0000-0000-0000-000000000001")
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    sid = await admin_login(client, admin.email, pw)
    resp = await client.get("/admin/audit")
    assert f'href="/admin/services/{svc.id}"' in resp.text
    assert "행위자서비스" in resp.text
    assert "외부 서비스" in resp.text


async def _seed_audit_rows(db, cipher):
    """검색/필터 검증용 로그 3건: USER 로그인, SERVICE 구독생성, 시스템 만료."""
    from app.services.audit import record_audit
    svc, _, _ = await create_service(db, cipher, name="검색대상서비스")
    user, _ = await create_user(db, role="SYSTEM_ADMIN", email="searchme@x.com")
    await record_audit(db, actor_type="USER", actor_user_id=user.id,
                       action="auth.login", detail={"note": "hello-detail"})
    await record_audit(db, actor_type="SERVICE", actor_service_id=svc.id,
                       action="subscription.create", target_type="subscription",
                       target_id="target-abc-123")
    await record_audit(db, actor_type="SYSTEM", action="subscription.expired")
    await db.commit()
    return svc, user


async def test_audit_action_filter(client, db, redis_client, cipher):
    await _seed_audit_rows(db, cipher)
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    resp = await client.get("/admin/audit?action=auth.login")
    tbody_start = resp.text.find("<tbody>")
    tbody = resp.text[tbody_start:]
    assert "로그인" in tbody
    assert "구독 생성" not in tbody
    # 필터 select에 활동 옵션 렌더
    assert '<select name="action"' in resp.text


async def test_audit_q_searches_actor_target_detail(client, db, redis_client, cipher):
    svc, user = await _seed_audit_rows(db, cipher)
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    # 행위자 이메일 like
    resp = await client.get("/admin/audit?q=searchme")
    tbody = resp.text[resp.text.find("<tbody>"):]
    assert "로그인" in tbody and "구독 생성" not in tbody
    # 행위자 서비스명 like
    resp = await client.get("/admin/audit?q=검색대상")
    tbody = resp.text[resp.text.find("<tbody>"):]
    assert "구독 생성" in tbody and "로그인" not in tbody
    # 대상 target_id like
    resp = await client.get("/admin/audit?q=target-abc")
    tbody = resp.text[resp.text.find("<tbody>"):]
    assert "구독 생성" in tbody
    # 상세 detail like (JSONB 텍스트 캐스팅)
    resp = await client.get("/admin/audit?q=hello-detail")
    tbody = resp.text[resp.text.find("<tbody>"):]
    assert "로그인" in tbody and "구독 생성" not in tbody


async def test_audit_export_xlsx(client, db, redis_client, cipher):
    """현재 필터가 적용된 감사로그를 xlsx로 다운로드."""
    from io import BytesIO
    from openpyxl import load_workbook
    await _seed_audit_rows(db, cipher)  # USER 로그인 + SERVICE 구독생성 + SYSTEM 만료 3건
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)

    resp = await client.get("/admin/audit/export.xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert 'attachment; filename="audit-log-' in resp.headers["content-disposition"]
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["시각", "행위자", "활동", "대상", "상세", "IP"]
    # 시드 3건 + 헤더 = 4행 이상 (admin 로그인 감사 로그가 더 있을 수 있음)
    assert ws.max_row >= 4
    actors = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert any(a and "외부 서비스 (검색대상서비스)" in a for a in actors)


async def test_audit_export_applies_filters(client, db, redis_client, cipher):
    """필터/검색이 적용된 결과만 내려받는다."""
    from io import BytesIO
    from openpyxl import load_workbook
    await _seed_audit_rows(db, cipher)
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)

    resp = await client.get("/admin/audit/export.xlsx?action=subscription.create")
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    actions = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
    assert actions == ["구독 생성"]


async def test_audit_page_has_export_button(client, db, redis_client, cipher):
    await _seed_audit_rows(db, cipher)
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    resp = await client.get("/admin/audit?action=auth.login")
    # 현재 쿼리스트링이 export 링크에 유지된다
    assert "/admin/audit/export.xlsx?" in resp.text
    assert "action=auth.login" in resp.text


async def test_subscription_service_filter_and_expiry_sort(client, db, redis_client, cipher):
    from datetime import timedelta
    from app.core.clock import utcnow
    svc_a, _, _ = await create_service(db, cipher, name="filt-a")
    svc_b, _, _ = await create_service(db, cipher, name="filt-b")
    plan_a = await create_plan(db, svc_a)
    plan_b = await create_plan(db, svc_b)
    await create_subscription(db, cipher, svc_a, plan_a, external_user_id="in-a",
                              period_end=utcnow() + timedelta(days=5))
    await create_subscription(db, cipher, svc_b, plan_b, external_user_id="in-b",
                              period_end=utcnow() + timedelta(days=20))
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)

    # 서비스 필터 드롭다운 노출(전체 + 서비스명)
    html = (await client.get("/admin/subscriptions")).text
    assert "전체 서비스" in html and "filt-a" in html and "filt-b" in html
    # 서비스 A로 필터 → A만
    only_a = (await client.get(f"/admin/subscriptions?service_id={svc_a.id}")).text
    assert "in-a" in only_a and "in-b" not in only_a
    # 만료일 오름차순 정렬 → in-a(5일)가 in-b(20일)보다 먼저
    asc = (await client.get(
        "/admin/subscriptions?sort=current_period_end&dir=asc")).text
    assert asc.index("in-a") < asc.index("in-b")


async def test_audit_purge_deletes_only_before_date(client, db, redis_client, cipher):
    """기준일 이전 로그만 삭제하고, 삭제 행위를 감사 기록한다."""
    from datetime import datetime, timezone
    from sqlalchemy import select as sa_select
    from app.models import AuditLog
    from app.services.audit import record_audit
    await record_audit(db, actor_type="SYSTEM", action="old.entry")
    await record_audit(db, actor_type="SYSTEM", action="new.entry")
    await db.commit()
    # old.entry를 과거로 보낸다 (created_at은 server_default라 직접 update)
    old = await db.scalar(sa_select(AuditLog).where(AuditLog.action == "old.entry"))
    old.created_at = datetime(2020, 1, 1, tzinfo=timezone.utc)
    await db.commit()

    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    admin_id = admin.id  # expire_all 이후 PK 접근 시 MissingGreenlet 방지
    sid = await admin_login(client, admin.email, pw)
    csrf = await get_csrf(redis_client, sid)
    resp = await client.post("/admin/audit/purge",
                             data={"csrf_token": csrf, "before": "2021-01-01"})
    assert resp.status_code == 303
    assert "flash=" in resp.headers["location"]

    db.expire_all()
    assert await db.scalar(sa_select(AuditLog).where(
        AuditLog.action == "old.entry")) is None        # 기준일 이전 → 삭제
    assert await db.scalar(sa_select(AuditLog).where(
        AuditLog.action == "new.entry")) is not None    # 이후 → 보존
    purge_log = await db.scalar(sa_select(AuditLog).where(
        AuditLog.action == "audit.purge"))
    assert purge_log is not None
    assert purge_log.detail["before"] == "2021-01-01"
    assert purge_log.detail["deleted_count"] == 1
    assert purge_log.actor_user_id == admin_id


async def test_audit_purge_invalid_date_shows_error(client, db, redis_client, cipher):
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    sid = await admin_login(client, admin.email, pw)
    csrf = await get_csrf(redis_client, sid)
    resp = await client.post("/admin/audit/purge",
                             data={"csrf_token": csrf, "before": "not-a-date"})
    assert resp.status_code == 303
    assert "flash_type=error" in resp.headers["location"]


async def test_audit_page_has_purge_form(client, db, redis_client, cipher):
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    resp = await client.get("/admin/audit")
    assert "/admin/audit/purge" in resp.text
    assert 'type="date"' in resp.text
    assert "data-confirm" in resp.text


async def test_audit_purge_rejects_future_date(client, db, redis_client, cipher):
    """미래 기준일 거부 — '과거 데이터 삭제' 취지."""
    from sqlalchemy import select as sa_select
    from app.models import AuditLog
    from app.services.audit import record_audit
    await record_audit(db, actor_type="SYSTEM", action="keep.entry")
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    sid = await admin_login(client, admin.email, pw)
    csrf = await get_csrf(redis_client, sid)
    resp = await client.post("/admin/audit/purge",
                             data={"csrf_token": csrf, "before": "2099-12-31"})
    assert resp.status_code == 303
    assert "flash_type=error" in resp.headers["location"]
    db.expire_all()
    assert await db.scalar(sa_select(AuditLog).where(
        AuditLog.action == "keep.entry")) is not None  # 아무것도 삭제되지 않음


async def test_audit_purge_requires_csrf(client, db, redis_client, cipher):
    # validate_csrf가 PermissionDeniedError(http_status=403)를 던지고
    # register_error_handlers의 DomainError 핸들러가 JSON 403을 반환한다.
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    resp = await client.post("/admin/audit/purge",
                             data={"csrf_token": "wrong", "before": "2021-01-01"})
    assert resp.status_code == 403


async def test_audit_purge_empty_date_shows_error(client, db, redis_client, cipher):
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    sid = await admin_login(client, admin.email, pw)
    csrf = await get_csrf(redis_client, sid)
    resp = await client.post("/admin/audit/purge",
                             data={"csrf_token": csrf, "before": ""})
    assert resp.status_code == 303
    assert "flash_type=error" in resp.headers["location"]


async def test_payments_date_range_filter(client, db, redis_client, cipher):
    """결제일(requested_at) 시작~끝 범위 필터 (기존 month 필터 대체)."""
    from datetime import datetime, timezone
    from app.models import Payment
    svc, _, _ = await create_service(db, cipher, name="pay-range-svc")
    plan = await create_plan(db, svc)
    sub = await create_subscription(db, cipher, svc, plan, external_user_id="pr-user")
    db.add(Payment(subscription_id=sub.id, order_id="pr-old", amount=1000,
                   payment_type="FIRST", status="DONE", idempotency_key="pr-old",
                   requested_at=datetime(2025, 3, 10, tzinfo=timezone.utc),
                   service_id=sub.service_id, external_user_id=sub.external_user_id))
    db.add(Payment(subscription_id=sub.id, order_id="pr-new", amount=2000,
                   payment_type="RENEWAL", status="DONE", idempotency_key="pr-new",
                   requested_at=datetime(2025, 4, 10, tzinfo=timezone.utc),
                   service_id=sub.service_id, external_user_id=sub.external_user_id))
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)

    resp = await client.get("/admin/payments?from=2025-03-01&to=2025-03-31")
    assert "pr-old" in resp.text and "pr-new" not in resp.text
    # 형식 오류는 무시(전체)
    resp = await client.get("/admin/payments?from=bogus")
    assert "pr-old" in resp.text and "pr-new" in resp.text
    # month 파라미터는 더 이상 동작하지 않음(전체 표시)
    resp = await client.get("/admin/payments?month=2025-03")
    assert "pr-old" in resp.text and "pr-new" in resp.text


async def test_subscriptions_date_range_filter(client, db, redis_client, cipher):
    """구독일(created_at) 시작~끝 범위 필터 — 경계 포함."""
    from datetime import datetime, timezone
    from sqlalchemy import update as sa_update
    from app.models import Subscription
    svc, _, _ = await create_service(db, cipher, name="sub-range-svc")
    plan = await create_plan(db, svc)
    s_in = await create_subscription(db, cipher, svc, plan, external_user_id="rng-in")
    s_out = await create_subscription(db, cipher, svc, plan, external_user_id="rng-out")
    # created_at은 server_default — 직접 update로 과거 날짜 부여
    await db.execute(sa_update(Subscription).where(Subscription.id == s_in.id)
                     .values(created_at=datetime(2026, 2, 15, tzinfo=timezone.utc)))
    await db.execute(sa_update(Subscription).where(Subscription.id == s_out.id)
                     .values(created_at=datetime(2026, 3, 5, tzinfo=timezone.utc)))
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)

    resp = await client.get("/admin/subscriptions?from=2026-02-01&to=2026-02-28")
    assert "rng-in" in resp.text and "rng-out" not in resp.text
    # 한쪽만 입력(열린 범위)
    resp = await client.get("/admin/subscriptions?from=2026-03-01")
    assert "rng-out" in resp.text and "rng-in" not in resp.text
    # date input 렌더
    assert 'type="date"' in resp.text


async def test_payments_kind_and_service_filter(client, db, redis_client, cipher):
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus
    svc_a, _, _ = await create_service(db, cipher, name="결제구분A")
    svc_b, _, _ = await create_service(db, cipher, name="결제구분B")
    plan = await create_plan(db, svc_a)
    sub = await create_subscription(db, cipher, svc_a, plan, external_user_id="sub-user")
    db.add(Payment(subscription_id=sub.id, service_id=svc_a.id, external_user_id="sub-user",
                   order_id="kind-sub", amount=1000, payment_type="RENEWAL",
                   kind=PaymentKind.SUBSCRIPTION, status=PaymentStatus.DONE,
                   idempotency_key="kind-sub", requested_at=utcnow()))
    db.add(Payment(subscription_id=None, service_id=svc_b.id, external_user_id="oo-user",
                   order_id="kind-oo", amount=2000, payment_type="ONE_OFF",
                   kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                   idempotency_key="kind-oo", requested_at=utcnow()))
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)

    def tbody(h): return h[h.find("<tbody>"):]
    body = tbody((await client.get("/admin/payments?kind=ONE_OFF")).text)
    assert "kind-oo" in body and "kind-sub" not in body
    body = tbody((await client.get(f"/admin/payments?service_id={svc_a.id}")).text)
    assert "kind-sub" in body and "kind-oo" not in body
    html = (await client.get("/admin/payments")).text
    assert 'name="kind"' in html and 'name="service_id"' in html
    assert "oo-user" in html        # 단건(구독 없음) 사용자 표시


async def test_payments_filter_order(client, db, redis_client, cipher):
    """필터 순서: 서비스 → 요금제 → 종류 → 상태 → 기간."""
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    html = (await client.get("/admin/payments")).text
    i_service = html.find('name="service_id"')
    i_plan = html.find('name="plan_name"')
    i_kind = html.find('name="kind"')
    i_status = html.find('name="status"')
    i_from = html.find('name="from"')
    assert -1 < i_service < i_plan < i_kind < i_status < i_from


async def test_subscriptions_plan_filter(client, db, redis_client, cipher):
    svc, _, _ = await create_service(db, cipher, name="요금제필터서비스")
    p_basic = await create_plan(db, svc, name="베이직플랜")
    p_pro = await create_plan(db, svc, name="프로플랜")
    await create_subscription(db, cipher, svc, p_basic, external_user_id="sub-basic")
    await create_subscription(db, cipher, svc, p_pro, external_user_id="sub-pro")
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    def tbody(h): return h[h.find("<tbody>"):]
    body = tbody((await client.get("/admin/subscriptions?plan_name=베이직플랜")).text)
    assert "sub-basic" in body and "sub-pro" not in body
    # 요금제 select + 서비스 선택 시 그 서비스 요금제 노출
    html = (await client.get(f"/admin/subscriptions?service_id={svc.id}")).text
    assert 'name="plan_name"' in html and "베이직플랜" in html and "프로플랜" in html


async def test_payments_plan_filter_subscription_only(client, db, redis_client, cipher):
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc, _, _ = await create_service(db, cipher, name="결제요금제서비스")
    plan = await create_plan(db, svc, name="요금제P")
    sub = await create_subscription(db, cipher, svc, plan, external_user_id="pf-sub")
    db.add(Payment(subscription_id=sub.id, service_id=svc.id, external_user_id="pf-sub",
                   order_id="pf-sub-pay", amount=1000, payment_type=PaymentType.RENEWAL,
                   kind=PaymentKind.SUBSCRIPTION, status=PaymentStatus.DONE,
                   idempotency_key="pf-sub-pay", requested_at=utcnow()))
    db.add(Payment(subscription_id=None, service_id=svc.id, external_user_id="pf-oo",
                   order_id="pf-oo-pay", amount=2000, payment_type=PaymentType.ONE_OFF,
                   kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                   idempotency_key="pf-oo-pay", requested_at=utcnow()))
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    def tbody(h): return h[h.find("<tbody>"):]
    body = tbody((await client.get("/admin/payments?plan_name=요금제P")).text)
    assert "pf-sub-pay" in body and "pf-oo-pay" not in body   # 일반결제 제외
    html = (await client.get(f"/admin/payments?service_id={svc.id}")).text
    assert 'name="plan_name"' in html and "요금제P" in html
    # 요금제 미선택 시 일반결제도 정상 포함(outerjoin 회귀)
    all_body = tbody((await client.get("/admin/payments")).text)
    assert "pf-sub-pay" in all_body and "pf-oo-pay" in all_body


async def test_payment_kind_type_badges(client, db, redis_client, cipher):
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc, _, _ = await create_service(db, cipher)
    now = utcnow()
    db.add(Payment(subscription_id=None, service_id=svc.id, external_user_id="b-oo",
                   order_id="badge-oo", amount=1000, payment_type=PaymentType.ONE_OFF,
                   kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                   idempotency_key="badge-oo", requested_at=now, approved_at=now))
    await db.commit()
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    html = (await client.get("/admin/payments")).text
    assert "badge-ONE_OFF" in html        # 종류 색 배지
    assert "badge badge-ONE_OFF" in html   # 유형도 배지(payment_type=ONE_OFF)
    # 정산 서비스별 모드에도 종류/유형 배지 적용
    s_html = (await client.get(
        f"/admin/settlement?from={now:%Y-%m-01}&to={now:%Y-%m-%d}&service_id={svc.id}")).text
    assert "badge badge-ONE_OFF" in s_html


async def test_payment_detail_page_and_scope(client, db, redis_client, cipher):
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    import uuid as _uuid
    svc_a, _, _ = await create_service(db, cipher, name="상세A")
    svc_b, _, _ = await create_service(db, cipher, name="상세B")
    p = Payment(subscription_id=None, service_id=svc_a.id, external_user_id="det-u",
                order_id="det-pay", amount=4200, payment_type=PaymentType.ONE_OFF,
                kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                idempotency_key="det-pay", requested_at=utcnow(), approved_at=utcnow())
    db.add(p); await db.commit(); await db.refresh(p)
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    html = (await client.get(f"/admin/payments/{p.id}")).text
    assert "det-pay" in html and "4,200" in html and "상세A" in html
    # 매니저 스코프: 타 서비스 결제 404
    mgr, pw2 = await create_user(db, role="SERVICE_MANAGER", service_id=svc_b.id)
    await admin_login(client, mgr.email, pw2)
    resp = await client.get(f"/admin/payments/{p.id}")
    assert resp.status_code == 404
    # 미존재 결제 → 404 (admin으로 재로그인 후 scope 무관 검증)
    await admin_login(client, admin.email, pw)
    resp_missing = await client.get(f"/admin/payments/{_uuid.uuid4()}")
    assert resp_missing.status_code == 404


async def test_external_cancel_api(client, db, redis_client, cipher, fake_toss):
    """외부 API: POST /api/v1/payments/{order_id}/cancel → 200, status CANCELED.

    DONE ONE_OFF 결제를 시드하고 HMAC 서명 헬퍼로 취소 요청. 취소 허용 서비스 기본값.
    """
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc, api_key, hmac_secret = await create_service(db, cipher, name="api-cancel-svc")
    p = Payment(subscription_id=None, service_id=svc.id, external_user_id="api-cancel-u",
                order_id="api-cancel-ord-1", amount=5000, payment_type=PaymentType.ONE_OFF,
                kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                idempotency_key="api-cancel-ik-1",
                toss_payment_key="pay_api-cancel-ord-1",
                requested_at=utcnow())
    db.add(p)
    await db.commit()
    # HMAC 서명 요청: reason 기본값 사용(빈 바디 허용)
    resp = await api_request(client, "POST",
                             f"/api/v1/payments/api-cancel-ord-1/cancel",
                             api_key, hmac_secret,
                             json_body={"reason": "테스트 취소"})
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["status"] == "CANCELED"
    assert data["order_id"] == "api-cancel-ord-1"
    # 토스 취소 호출 확인
    assert any(c["payment_key"] == "pay_api-cancel-ord-1" for c in fake_toss.canceled)


async def test_external_cancel_api_disabled_service(client, db, redis_client,
                                                     cipher, fake_toss):
    """외부 API: cancellation_enabled=False 서비스 → 취소 시도 시 4xx(취소 불가).

    DONE ONE_OFF 결제를 시드하고 HMAC 서명 헬퍼로 취소 요청.
    서비스 정책(cancellation_enabled=False)으로 취소 불가 → 4xx 응답 검증.
    """
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc, api_key, hmac_secret = await create_service(db, cipher, name="api-cancel-disabled-svc")
    svc.cancellation_enabled = False
    await db.commit()

    p = Payment(subscription_id=None, service_id=svc.id, external_user_id="api-disabled-u",
                order_id="api-disabled-ord-1", amount=5000, payment_type=PaymentType.ONE_OFF,
                kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                idempotency_key="api-disabled-ik-1",
                toss_payment_key="pay_api-disabled-ord-1",
                requested_at=utcnow())
    db.add(p)
    await db.commit()

    # HMAC 서명 요청: cancellation_enabled=False 서비스이므로 취소 불가 → 4xx
    resp = await api_request(client, "POST",
                             "/api/v1/payments/api-disabled-ord-1/cancel",
                             api_key, hmac_secret,
                             json_body={"reason": "취소 불가 테스트"})
    assert resp.status_code >= 400, f"Expected 4xx, got {resp.status_code}: {resp.text}"
    # 토스 취소 API가 호출되지 않았음을 확인
    assert not any(c["payment_key"] == "pay_api-disabled-ord-1"
                   for c in fake_toss.canceled)


async def test_admin_payment_cancel_button_and_action(client, db, redis_client,
                                                       cipher, fake_toss):
    """Admin 결제상세: 취소 버튼 노출 + POST 전액취소 → CANCELED(수수료 없이 전액 환불).

    어드민 취소는 항상 허용이라 cancellation_enabled=False 서비스 결제도 취소 카드가 보인다.
    서비스 수수료율이 있어도 어드민 취소는 무시하고 전액 환불한다.
    """
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc_ok, _, _ = await create_service(db, cipher, name="cancel-ok-svc")
    svc_ok.cancellation_fee_percent = 10  # 수수료율 있어도 어드민 취소는 무시(전액 환불)
    svc_no, _, _ = await create_service(db, cipher, name="cancel-no-svc")
    svc_no.cancellation_enabled = False   # 게이트 꺼져 있어도 어드민은 항상 취소 가능
    await db.commit()

    p_ok = Payment(subscription_id=None, service_id=svc_ok.id, external_user_id="c-ok-u",
                   order_id="cancel-ok-ord", amount=3000, payment_type=PaymentType.ONE_OFF,
                   kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                   idempotency_key="cancel-ok-ik",
                   toss_payment_key="pay_cancel-ok-ord",
                   requested_at=utcnow())
    p_no = Payment(subscription_id=None, service_id=svc_no.id, external_user_id="c-no-u",
                   order_id="cancel-no-ord", amount=3000, payment_type=PaymentType.ONE_OFF,
                   kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                   idempotency_key="cancel-no-ik",
                   toss_payment_key="pay_cancel-no-ord",
                   requested_at=utcnow())
    db.add(p_ok)
    db.add(p_no)
    await db.commit()
    await db.refresh(p_ok)
    await db.refresh(p_no)

    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    session_id = await admin_login(client, admin.email, pw)
    csrf = await get_csrf(redis_client, session_id)

    # 결제 상세: '결제 취소' 카드 노출
    html_ok = (await client.get(f"/admin/payments/{p_ok.id}")).text
    assert "결제 취소" in html_ok

    # 게이트 꺼진 서비스도 어드민 취소 카드가 보인다(항상 허용)
    html_no = (await client.get(f"/admin/payments/{p_no.id}")).text
    assert "결제 취소" in html_no
    assert "취소 불가" not in html_no

    # Admin POST 전액취소(cancel_amount 미지정) → 303 redirect, 상태 CANCELED
    resp = await client.post(f"/admin/payments/{p_ok.id}/cancel",
                             data={"csrf_token": csrf})
    assert resp.status_code == 303
    assert f"/admin/payments/{p_ok.id}" in resp.headers["location"]
    await db.refresh(p_ok)
    assert p_ok.status == "CANCELED"
    assert p_ok.canceled_amount == 3000 and not p_ok.cancel_fee  # 수수료 없이 전액 환불

    # 취소 완료 후 결제상세에 취소·환불 내역 표시(전액 환불 3,000)
    html_canceled = (await client.get(f"/admin/payments/{p_ok.id}")).text
    assert "취소·환불 내역" in html_canceled
    assert "3,000" in html_canceled

    # 결제 목록에 취소 건이 '취소' 상태 + 환불 정보로 표시
    list_html = (await client.get("/admin/payments")).text
    assert "cancel-ok-ord" in list_html
    assert "취소" in list_html and "3,000" in list_html      # 한글 상태 + 환불액

    # 서비스 상세 일반결제 내역에도 환불 정보 표시
    svc_html = (await client.get(f"/admin/services/{svc_ok.id}")).text
    assert "cancel-ok-ord" in svc_html and "3,000" in svc_html


async def test_admin_payment_cancel_requires_csrf(client, db, redis_client, cipher):
    """Admin 취소 POST에 잘못된 CSRF 토큰이면 403."""
    from app.core.clock import utcnow
    from app.models import Payment, PaymentKind, PaymentStatus, PaymentType
    svc, _, _ = await create_service(db, cipher, name="csrf-cancel-svc")
    p = Payment(subscription_id=None, service_id=svc.id, external_user_id="csrf-u",
                order_id="csrf-cancel-ord", amount=1000, payment_type=PaymentType.ONE_OFF,
                kind=PaymentKind.ONE_OFF, status=PaymentStatus.DONE,
                idempotency_key="csrf-cancel-ik",
                toss_payment_key="pay_csrf-cancel-ord",
                requested_at=utcnow())
    db.add(p)
    await db.commit()
    await db.refresh(p)
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    resp = await client.post(f"/admin/payments/{p.id}/cancel",
                             data={"csrf_token": "wrong-token"})
    assert resp.status_code == 403


async def test_saved_modal_trigger_on_write_success(client, db, redis_client, cipher):
    """저장 성공 후 follow_redirects=True로 도달한 페이지에 body[data-saved]가 존재한다.

    대표 저장 액션: POST /admin/settings/retry → /admin/settings?saved=... →
    렌더된 HTML에 data-saved 속성이 포함되어야 완료 모달(✓)이 뜬다.
    """
    from app.services import app_settings

    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    session_id = await admin_login(client, admin.email, pw)
    csrf = await get_csrf(redis_client, session_id)

    # 저장 액션 후 follow_redirects=True로 최종 페이지 확인
    resp = await client.post(
        "/admin/settings/retry",
        data={
            "csrf_token": csrf,
            "retry_limit": "3",
            "retry_interval_hours": "4",
            "suspended_grace_days": "7",
        },
        follow_redirects=True,
    )
    assert resp.status_code == 200
    # 완료 모달 트리거: body 태그에 data-saved 속성이 포함되어야 함
    assert 'data-saved="' in resp.text, (
        "저장 성공 후 body[data-saved] 속성이 없습니다 (완료 모달 트리거 불가)"
    )


async def test_service_detail_events_section_shows_audit_detail(
        client, db, redis_client, cipher):
    """서비스 상세 하단 '이벤트' 섹션에 변경 동작이 상세(변경 전→후)와 함께 노출(요청 015).

    상태 변경·IP 갱신을 수행한 뒤 서비스 상세를 열어 이벤트 활동·상세 텍스트를 검증한다.
    """
    from app.services import registry
    svc, _, _ = await create_service(db, cipher, name="이벤트서비스",
                                     allowed_ips=["10.0.0.1"])
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    await admin_login(client, admin.email, pw)
    # 비활성화(상태 변경) + IP 갱신 → 감사 detail에 old/new 기록
    await registry.set_service_status(db, svc.id, "INACTIVE", actor_user_id=admin.id)
    await registry.update_allowed_ips(db, svc.id, ["10.0.0.1", "10.0.0.2"],
                                      actor_user_id=admin.id)
    html = (await client.get(f"/admin/services/{svc.id}")).text
    assert "이벤트" in html                                   # 이벤트 섹션
    assert "서비스 상태 변경" in html and "허용 IP 변경" in html  # 활동 라벨
    assert "ACTIVE → INACTIVE" in html                       # 상태 변경 전→후
    assert "10.0.0.1 → 10.0.0.1, 10.0.0.2" in html           # IP 변경 전→후


async def test_subscription_extend_via_admin(client, db, redis_client, cipher):
    """구독 상세에서 만료일 연장 POST → 303 + status=EXTENDED·만료일/다음결제=입력일, 목록에 '연장처리'."""
    from datetime import date, timedelta
    from app.core.clock import utcnow
    svc, _, _ = await create_service(db, cipher, name="연장서비스")
    plan = await create_plan(db, svc)
    sub = await create_subscription(db, cipher, svc, plan, external_user_id="ext-e2e",
                                    status="ACTIVE")
    admin, pw = await create_user(db, role="SYSTEM_ADMIN")
    session_id = await admin_login(client, admin.email, pw)
    csrf = await get_csrf(redis_client, session_id)

    target = (utcnow() + timedelta(days=90)).date().isoformat()
    resp = await client.post(f"/admin/subscriptions/{sub.id}/extend",
                             data={"csrf_token": csrf, "new_end": target})
    assert resp.status_code == 303
    await db.refresh(sub)
    assert sub.status == "EXTENDED"
    assert sub.current_period_end.date() == date.fromisoformat(target)
    assert sub.next_billing_at.date() == date.fromisoformat(target)

    # 목록·상세에 한글 '연장처리' 노출
    list_html = (await client.get("/admin/subscriptions")).text
    assert "연장처리" in list_html
    detail_html = (await client.get(f"/admin/subscriptions/{sub.id}")).text
    assert "연장처리" in detail_html
    # 구독 상세에 구독 ID(KEY) + 연장 이력 노출
    assert "구독 ID(KEY)" in detail_html and str(sub.id) in detail_html
    assert "만료일 연장 이력" in detail_html and target in detail_html  # 변경 후 만료일(YYYY-MM-DD)
