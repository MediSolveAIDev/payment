from io import BytesIO

from openpyxl import load_workbook

from app.admin.export import XLSX_MEDIA, xlsx_response, xlsx_safe


def test_xlsx_safe_guards_formula():
    assert xlsx_safe("=SUM(A1)") == "'=SUM(A1)"
    assert xlsx_safe("+1") == "'+1"
    assert xlsx_safe("-cmd") == "'-cmd"
    assert xlsx_safe("@x") == "'@x"
    assert xlsx_safe("") == ""
    assert xlsx_safe("정상") == "정상"
    assert xlsx_safe(1000) == 1000           # 숫자는 그대로


def test_xlsx_response_headers_and_content():
    resp = xlsx_response("services", ["이름", "값"], [["x", 1], ["=y", 2]])
    assert resp.media_type == XLSX_MEDIA
    cd = resp.headers["content-disposition"]
    assert "attachment" in cd and "services-" in cd and ".xlsx" in cd
    wb = load_workbook(BytesIO(resp.body))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["이름", "값"]
    assert ws[3][0].value == "'=y"           # 수식 방어
    assert ws[2][1].value == 1


def test_xlsx_response_korean_filename():
    resp = xlsx_response("서비스-구독", ["A"], [["x"]])
    cd = resp.headers["content-disposition"]
    assert "filename*=UTF-8''" in cd          # RFC 5987
    assert "attachment" in cd
